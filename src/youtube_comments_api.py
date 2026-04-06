import json
import os
import time
from pathlib import Path
import urllib.parse
import urllib.error
import urllib.request
from typing import Any, Callable, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def normalize_reply(reply_item: Dict[str, Any]) -> Dict[str, Any]:
    snippet = reply_item["snippet"]
    return {
        "comment_id": reply_item["id"],
        "parent_id": snippet.get("parentId"),
        "author": snippet.get("authorDisplayName"),
        "published_at": snippet.get("publishedAt"),
        "text": snippet.get("textDisplay"),
        "like_count": snippet.get("likeCount", 0),
    }


def normalize_thread(thread_item: Dict[str, Any]) -> Dict[str, Any]:
    snippet = thread_item["snippet"]
    top_level = snippet["topLevelComment"]["snippet"]
    embedded_replies: List[Dict[str, Any]] = [
        normalize_reply(reply) for reply in thread_item.get("replies", {}).get("comments", [])
    ]

    return {
        "thread_id": thread_item["id"],
        "comment_id": thread_item["snippet"]["topLevelComment"]["id"],
        "author": top_level.get("authorDisplayName"),
        "published_at": top_level.get("publishedAt"),
        "text": top_level.get("textDisplay"),
        "like_count": top_level.get("likeCount", 0),
        "reply_count": snippet.get("totalReplyCount", 0),
        "replies": embedded_replies,
    }


class YouTubeDataApiClient:
    def __init__(
        self,
        api_key: str,
        opener: Optional[Callable[..., Any]] = None,
        retries: int = 3,
        retry_sleep: float = 1.0,
    ):
        self.api_key = api_key
        self.opener = opener or urllib.request.urlopen
        self.retries = retries
        self.retry_sleep = retry_sleep

    def _request_json(self, endpoint: str, params: Dict[str, Any]) -> Dict[str, Any]:
        query = dict(params)
        query["key"] = self.api_key
        url = endpoint + "?" + urllib.parse.urlencode(query)
        attempt = 0
        while True:
            try:
                with self.opener(url, timeout=30) as response:
                    return json.load(response)
            except urllib.error.URLError:
                attempt += 1
                if attempt > self.retries:
                    raise
                time.sleep(self.retry_sleep)

    def list_comment_threads(
        self, video_id: str, page_token: Optional[str] = None, order: str = "relevance"
    ) -> Dict[str, Any]:
        params = {
            "part": "snippet,replies",
            "videoId": video_id,
            "maxResults": 100,
            "order": order,
            "textFormat": "plainText",
        }
        if page_token:
            params["pageToken"] = page_token
        return self._request_json("https://www.googleapis.com/youtube/v3/commentThreads", params)

    def list_comment_replies(self, parent_id: str, page_token: Optional[str] = None) -> Dict[str, Any]:
        params = {
            "part": "snippet",
            "parentId": parent_id,
            "maxResults": 100,
            "textFormat": "plainText",
        }
        if page_token:
            params["pageToken"] = page_token
        return self._request_json("https://www.googleapis.com/youtube/v3/comments", params)


class YouTubeCommentsExporter:
    def __init__(self, api_client: Any):
        self.api_client = api_client

    def _fetch_missing_replies(
        self, parent_comment_id: str, existing_replies: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        seen_ids = {reply["comment_id"] for reply in existing_replies}
        replies = list(existing_replies)
        page_token: Optional[str] = None

        while True:
            payload = self.api_client.list_comment_replies(parent_comment_id, page_token=page_token)
            for item in payload.get("items", []):
                normalized = normalize_reply(item)
                if normalized["comment_id"] not in seen_ids:
                    replies.append(normalized)
                    seen_ids.add(normalized["comment_id"])
            page_token = payload.get("nextPageToken")
            if not page_token:
                break

        replies.sort(key=lambda reply: (reply.get("published_at") or "", reply["comment_id"]))
        return replies

    def export_video_comments(self, video_id: str, order: str = "relevance") -> Dict[str, Any]:
        threads: List[Dict[str, Any]] = []
        page_token: Optional[str] = None

        while True:
            payload = self.api_client.list_comment_threads(video_id, page_token=page_token, order=order)
            for item in payload.get("items", []):
                normalized = normalize_thread(item)
                if normalized["reply_count"] > len(normalized["replies"]):
                    normalized["replies"] = self._fetch_missing_replies(
                        normalized["comment_id"], normalized["replies"]
                    )
                threads.append(normalized)
            page_token = payload.get("nextPageToken")
            if not page_token:
                break

        reply_count = sum(len(thread["replies"]) for thread in threads)
        return {
            "video_id": video_id,
            "summary": {
                "top_level_comment_count": len(threads),
                "reply_count": reply_count,
                "total_comment_count": len(threads) + reply_count,
            },
            "threads": threads,
        }


def extract_video_id(url_or_video_id: str) -> str:
    if "youtube.com" not in url_or_video_id and "youtu.be" not in url_or_video_id:
        return url_or_video_id

    parsed = urllib.parse.urlparse(url_or_video_id)
    if parsed.netloc.endswith("youtu.be"):
        return parsed.path.lstrip("/")

    query = urllib.parse.parse_qs(parsed.query)
    video_ids = query.get("v", [])
    if not video_ids:
        raise ValueError("Unable to extract video ID from URL")
    return video_ids[0]


def build_screenshot_match_report(export: Dict[str, Any]) -> Dict[str, Any]:
    expected_author = "@chinahamyku6583"
    expected_text_fragment = "A Thousand Miles of Rivers and Mountains"
    expected_reply_authors = ["@ireneserrano4570", "@vernicetan3089"]

    for thread in export.get("threads", []):
        text = thread.get("text") or ""
        if thread.get("author") == expected_author and expected_text_fragment in text:
            reply_authors = [reply.get("author") for reply in thread.get("replies", [])]
            return {
                "thread_found": True,
                "matched_thread": {
                    "thread_id": thread.get("thread_id"),
                    "author": thread.get("author"),
                    "published_at": thread.get("published_at"),
                    "like_count": thread.get("like_count"),
                    "reply_count": thread.get("reply_count"),
                    "text_preview": text[:240],
                },
                "matched_reply_count": len(thread.get("replies", [])),
                "matched_reply_authors": [
                    author for author in reply_authors if author in expected_reply_authors
                ],
                "all_reply_authors": reply_authors,
            }

    return {
        "thread_found": False,
        "matched_thread": None,
        "matched_reply_count": 0,
        "matched_reply_authors": [],
        "all_reply_authors": [],
    }


def _style_workbook_header(worksheet: Any) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font


def _set_widths(worksheet: Any, widths: Dict[int, float]) -> None:
    for column_index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def write_flat_excel_workbook(export: Dict[str, Any], output_path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FlatComments"
    worksheet.append(
        [
            "level",
            "thread_id",
            "comment_id",
            "parent_comment_id",
            "author",
            "published_at",
            "like_count",
            "reply_count",
            "text",
        ]
    )

    for thread in export.get("threads", []):
        worksheet.append(
            [
                "top",
                thread.get("thread_id"),
                thread.get("comment_id"),
                "",
                thread.get("author"),
                thread.get("published_at"),
                thread.get("like_count"),
                thread.get("reply_count"),
                thread.get("text"),
            ]
        )
        for reply in thread.get("replies", []):
            worksheet.append(
                [
                    "reply",
                    thread.get("thread_id"),
                    reply.get("comment_id"),
                    thread.get("comment_id"),
                    reply.get("author"),
                    reply.get("published_at"),
                    reply.get("like_count"),
                    "",
                    reply.get("text"),
                ]
            )

    _style_workbook_header(worksheet)
    worksheet.freeze_panes = "A2"
    _set_widths(
        worksheet,
        {1: 10, 2: 26, 3: 30, 4: 26, 5: 24, 6: 24, 7: 12, 8: 12, 9: 120},
    )

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()
    return output


def write_threaded_excel_workbook(export: Dict[str, Any], output_path: Path) -> Path:
    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "Summary"
    ws_threads = workbook.create_sheet("Threads")
    ws_replies = workbook.create_sheet("Replies")

    for row in [
        ["video_id", export.get("video_id")],
        ["order", export.get("order")],
        ["top_level_comment_count", export.get("summary", {}).get("top_level_comment_count", 0)],
        ["reply_count", export.get("summary", {}).get("reply_count", 0)],
        ["total_comment_count", export.get("summary", {}).get("total_comment_count", 0)],
    ]:
        ws_summary.append(row)
    _set_widths(ws_summary, {1: 24, 2: 24})

    ws_threads.append(
        ["thread_id", "comment_id", "author", "published_at", "like_count", "reply_count", "text"]
    )
    for thread in export.get("threads", []):
        ws_threads.append(
            [
                thread.get("thread_id"),
                thread.get("comment_id"),
                thread.get("author"),
                thread.get("published_at"),
                thread.get("like_count"),
                thread.get("reply_count"),
                thread.get("text"),
            ]
        )
    _style_workbook_header(ws_threads)
    ws_threads.freeze_panes = "A2"
    _set_widths(ws_threads, {1: 26, 2: 26, 3: 24, 4: 24, 5: 12, 6: 12, 7: 120})

    ws_replies.append(
        [
            "thread_id",
            "parent_comment_id",
            "reply_comment_id",
            "author",
            "published_at",
            "like_count",
            "text",
        ]
    )
    for thread in export.get("threads", []):
        for reply in thread.get("replies", []):
            ws_replies.append(
                [
                    thread.get("thread_id"),
                    thread.get("comment_id"),
                    reply.get("comment_id"),
                    reply.get("author"),
                    reply.get("published_at"),
                    reply.get("like_count"),
                    reply.get("text"),
                ]
            )
    _style_workbook_header(ws_replies)
    ws_replies.freeze_panes = "A2"
    _set_widths(ws_replies, {1: 26, 2: 26, 3: 30, 4: 24, 5: 24, 6: 12, 7: 120})

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    workbook.close()
    return output


def load_export_json(json_path: Path) -> Dict[str, Any]:
    return json.loads(Path(json_path).read_text(encoding="utf-8"))


def write_flat_excel_from_json(json_path: Path, output_path: Path) -> Path:
    export = load_export_json(json_path)
    return write_flat_excel_workbook(export, output_path)


def build_output_paths(output_dir: Path, video_id: str, order: str) -> Dict[str, Path]:
    stem = f"{video_id}.{order}.comments"
    return {
        "json": output_dir / f"{stem}.json",
        "compare": output_dir / f"{video_id}.{order}.compare.json",
        "threaded_excel": output_dir / f"{stem}.xlsx",
        "flat_excel": output_dir / f"{stem}.flat.xlsx",
    }


def export_video_to_path(
    api_key: str,
    url_or_video_id: str,
    output_path: str,
    compare_path: Optional[str] = None,
    order: str = "relevance",
) -> Dict[str, Any]:
    video_id = extract_video_id(url_or_video_id)
    exporter = YouTubeCommentsExporter(YouTubeDataApiClient(api_key))
    export = exporter.export_video_comments(video_id, order=order)
    export["order"] = order

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(export, ensure_ascii=False, indent=2), encoding="utf-8")

    if compare_path:
        compare_report = build_screenshot_match_report(export)
        compare_output = Path(compare_path)
        compare_output.parent.mkdir(parents=True, exist_ok=True)
        compare_output.write_text(json.dumps(compare_report, ensure_ascii=False, indent=2), encoding="utf-8")

    return export


def export_all_artifacts(
    api_key: str,
    url_or_video_id: str,
    output_dir: Path,
    order: str = "time",
) -> Dict[str, Any]:
    video_id = extract_video_id(url_or_video_id)
    paths = build_output_paths(Path(output_dir), video_id, order)
    export = export_video_to_path(
        api_key=api_key,
        url_or_video_id=url_or_video_id,
        output_path=str(paths["json"]),
        compare_path=str(paths["compare"]),
        order=order,
    )
    write_threaded_excel_workbook(export, paths["threaded_excel"])
    write_flat_excel_workbook(export, paths["flat_excel"])
    return {
        "video_id": video_id,
        "order": order,
        "summary": export["summary"],
        "paths": {name: str(path) for name, path in paths.items()},
    }


def main(argv: Optional[List[str]] = None) -> int:
    import argparse

    parser = argparse.ArgumentParser(description="Export YouTube comments through the official API")
    parser.add_argument("--url", required=True, help="YouTube watch URL or raw video ID")
    parser.add_argument("--output", required=True, help="Path to write the full export JSON")
    parser.add_argument("--compare-output", help="Optional path to write screenshot comparison JSON")
    parser.add_argument("--order", default="relevance", choices=["relevance", "time"])
    args = parser.parse_args(argv)

    api_key = os.environ.get("YOUTUBE_API_KEY")
    if not api_key:
        raise SystemExit("YOUTUBE_API_KEY is required")

    export = export_video_to_path(
        api_key=api_key,
        url_or_video_id=args.url,
        output_path=args.output,
        compare_path=args.compare_output,
        order=args.order,
    )
    print(
        json.dumps(
            {
                "video_id": export["video_id"],
                "order": export["order"],
                "summary": export["summary"],
                "output": args.output,
                "compare_output": args.compare_output,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
