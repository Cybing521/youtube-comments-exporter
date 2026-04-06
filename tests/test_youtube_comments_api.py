import json
import os
import subprocess
import sys
import tempfile
import unittest
import urllib.error
from pathlib import Path

from openpyxl import load_workbook

from src.youtube_comments_api import (
    YouTubeDataApiClient,
    YouTubeCommentsExporter,
    build_screenshot_match_report,
    extract_video_id,
    normalize_thread,
    write_flat_excel_workbook,
)


class NormalizeThreadTests(unittest.TestCase):
    def test_normalize_thread_keeps_requested_top_level_fields(self):
        item = {
            "id": "thread-1",
            "snippet": {
                "totalReplyCount": 2,
                "topLevelComment": {
                    "id": "comment-1",
                    "snippet": {
                        "authorDisplayName": "@author",
                        "publishedAt": "2026-04-01T00:00:00Z",
                        "textDisplay": "hello world",
                        "likeCount": 7,
                    },
                },
            },
            "replies": {
                "comments": [
                    {
                        "id": "reply-1",
                        "snippet": {
                            "authorDisplayName": "@reply-user",
                            "publishedAt": "2026-04-02T00:00:00Z",
                            "textDisplay": "reply text",
                            "likeCount": 3,
                            "parentId": "thread-1",
                        },
                    }
                ]
            },
        }

        normalized = normalize_thread(item)

        self.assertEqual("thread-1", normalized["thread_id"])
        self.assertEqual("@author", normalized["author"])
        self.assertEqual("2026-04-01T00:00:00Z", normalized["published_at"])
        self.assertEqual("hello world", normalized["text"])
        self.assertEqual(7, normalized["like_count"])
        self.assertEqual(2, normalized["reply_count"])
        self.assertEqual(1, len(normalized["replies"]))
        self.assertEqual("@reply-user", normalized["replies"][0]["author"])


class ExporterTests(unittest.TestCase):
    def test_exporter_paginates_threads_and_backfills_missing_replies(self):
        class FakeApiClient:
            def __init__(self):
                self.comments_calls = []

            def list_comment_threads(self, video_id, page_token=None, order="relevance"):
                if page_token is None:
                    return {
                        "nextPageToken": "page-2",
                        "items": [
                            {
                                "id": "thread-1",
                                "snippet": {
                                    "totalReplyCount": 2,
                                    "topLevelComment": {
                                        "id": "comment-1-parent",
                                        "snippet": {
                                            "authorDisplayName": "@author-1",
                                            "publishedAt": "2026-04-01T00:00:00Z",
                                            "textDisplay": "thread text 1",
                                            "likeCount": 10,
                                        },
                                    },
                                },
                                "replies": {
                                    "comments": [
                                        {
                                            "id": "reply-1",
                                            "snippet": {
                                                "authorDisplayName": "@reply-1",
                                                "publishedAt": "2026-04-01T00:05:00Z",
                                                "textDisplay": "reply one",
                                                "likeCount": 1,
                                                "parentId": "thread-1",
                                            },
                                        }
                                    ]
                                },
                            }
                        ],
                    }

                return {
                    "items": [
                        {
                            "id": "thread-2",
                            "snippet": {
                                    "totalReplyCount": 0,
                                    "topLevelComment": {
                                        "id": "comment-2-parent",
                                        "snippet": {
                                            "authorDisplayName": "@author-2",
                                            "publishedAt": "2026-04-02T00:00:00Z",
                                            "textDisplay": "thread text 2",
                                            "likeCount": 5,
                                    },
                                },
                            },
                        }
                    ]
                }

            def list_comment_replies(self, parent_id, page_token=None):
                self.comments_calls.append((parent_id, page_token))
                if page_token is None:
                    return {
                        "items": [
                            {
                                "id": "reply-2",
                                "snippet": {
                                    "authorDisplayName": "@reply-2",
                                    "publishedAt": "2026-04-01T00:06:00Z",
                                    "textDisplay": "reply two",
                                    "likeCount": 2,
                                    "parentId": "thread-1",
                                },
                            }
                        ]
                    }
                return {"items": []}

        exporter = YouTubeCommentsExporter(FakeApiClient())

        result = exporter.export_video_comments("video-1")

        self.assertEqual("video-1", result["video_id"])
        self.assertEqual(2, result["summary"]["top_level_comment_count"])
        self.assertEqual(2, result["summary"]["reply_count"])
        self.assertEqual(4, result["summary"]["total_comment_count"])
        self.assertEqual(["reply-1", "reply-2"], [r["comment_id"] for r in result["threads"][0]["replies"]])
        self.assertEqual([("comment-1-parent", None)], exporter.api_client.comments_calls)


class HelperTests(unittest.TestCase):
    def test_extract_video_id_from_watch_url(self):
        url = "https://www.youtube.com/watch?v=gtEROmL0NzQ&list=RDgtEROmL0NzQ&start_radio=1"
        self.assertEqual("gtEROmL0NzQ", extract_video_id(url))

    def test_build_screenshot_match_report_finds_thread_and_replies(self):
        export = {
            "video_id": "gtEROmL0NzQ",
            "threads": [
                {
                    "thread_id": "thread-top",
                    "author": "@chinahamyku6583",
                    "published_at": "2022-01-01T00:00:00Z",
                    "text": (
                        'This dance originated from one of the top ten famous paintings in China, '
                        '"A Thousand Miles of Rivers and Mountains".'
                    ),
                    "like_count": 3100,
                    "reply_count": 2,
                    "replies": [
                        {
                            "comment_id": "reply-a",
                            "author": "@ireneserrano4570",
                            "published_at": "2022-01-02T00:00:00Z",
                            "text": "dude... stop embarrassing yourself.",
                            "like_count": 69,
                            "parent_id": "thread-top",
                        },
                        {
                            "comment_id": "reply-b",
                            "author": "@vernicetan3089",
                            "published_at": "2022-01-03T00:00:00Z",
                            "text": "Thank you",
                            "like_count": 54,
                            "parent_id": "thread-top",
                        },
                    ],
                }
            ],
        }

        report = build_screenshot_match_report(export)

        self.assertTrue(report["thread_found"])
        self.assertEqual("@chinahamyku6583", report["matched_thread"]["author"])
        self.assertEqual(2, report["matched_reply_count"])
        self.assertEqual(["@ireneserrano4570", "@vernicetan3089"], report["matched_reply_authors"])


class ScriptTests(unittest.TestCase):
    def test_cli_script_runs_help_without_import_errors(self):
        project_root = Path(__file__).resolve().parent.parent
        script_path = project_root / "scripts" / "export_video_comments.py"

        completed = subprocess.run(
            [sys.executable, str(script_path), "--help"],
            cwd=project_root,
            capture_output=True,
            text=True,
        )

        self.assertEqual(0, completed.returncode, completed.stderr)
        self.assertIn("Export YouTube comments", completed.stdout)


class ApiClientTests(unittest.TestCase):
    def test_request_json_retries_after_transient_urlerror(self):
        class FakeResponse:
            def __init__(self, payload):
                self.payload = payload

            def __enter__(self):
                from io import StringIO

                self.stream = StringIO(json.dumps(self.payload))
                return self.stream

            def __exit__(self, exc_type, exc, tb):
                self.stream.close()
                return False

        calls = {"count": 0}

        def flaky_opener(url, timeout=30):
            calls["count"] += 1
            if calls["count"] == 1:
                raise urllib.error.URLError("temporary ssl eof")
            return FakeResponse({"items": [], "nextPageToken": None})

        client = YouTubeDataApiClient("test-key", opener=flaky_opener, retries=2, retry_sleep=0)

        payload = client.list_comment_replies("thread-1")

        self.assertEqual(2, calls["count"])
        self.assertEqual([], payload["items"])


class ExcelExportTests(unittest.TestCase):
    def test_write_flat_excel_workbook_creates_one_row_per_comment_or_reply(self):
        export = {
            "video_id": "video-1",
            "order": "time",
            "summary": {
                "top_level_comment_count": 1,
                "reply_count": 2,
                "total_comment_count": 3,
            },
            "threads": [
                {
                    "thread_id": "thread-1",
                    "comment_id": "comment-1",
                    "author": "@author",
                    "published_at": "2026-04-01T00:00:00Z",
                    "text": "top text",
                    "like_count": 10,
                    "reply_count": 2,
                    "replies": [
                        {
                            "comment_id": "reply-1",
                            "parent_id": "comment-1",
                            "author": "@reply-a",
                            "published_at": "2026-04-01T00:01:00Z",
                            "text": "reply a",
                            "like_count": 1,
                        },
                        {
                            "comment_id": "reply-2",
                            "parent_id": "comment-1",
                            "author": "@reply-b",
                            "published_at": "2026-04-01T00:02:00Z",
                            "text": "reply b",
                            "like_count": 2,
                        },
                    ],
                }
            ],
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "flat.xlsx"
            write_flat_excel_workbook(export, path)
            wb = load_workbook(path, read_only=True)

            self.assertEqual(["FlatComments"], wb.sheetnames)
            ws = wb["FlatComments"]
            self.assertEqual(4, ws.max_row)
            self.assertEqual("level", ws["A1"].value)
            self.assertEqual("top", ws["A2"].value)
            self.assertEqual("reply", ws["A3"].value)
            self.assertEqual("reply", ws["A4"].value)
            self.assertEqual("comment-1", ws["C2"].value)
            self.assertEqual("reply-1", ws["C3"].value)
            wb.close()


class GuiSmokeTests(unittest.TestCase):
    def test_mac_app_module_builds_main_window(self):
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        from src.mac_app import build_app, build_main_window

        app = build_app([])
        window = build_main_window()

        self.assertIsNotNone(app)
        self.assertEqual("YouTube 评论导出器", window.windowTitle())
        self.assertEqual("导出全部", window.export_button.text())
        self.assertEqual("从 JSON 生成扁平 Excel", window.flat_button.text())
        self.assertEqual("粘贴 YouTube 视频链接", window.url_input.placeholderText())
        self.assertGreaterEqual(window.url_input.minimumHeight(), 52)
        self.assertGreaterEqual(window.api_key_input.minimumHeight(), 52)
        self.assertEqual("运行日志", window.activity_log.placeholderText())
        self.assertTrue(hasattr(window, "entry_animations"))
        window.close()

    def test_export_all_uses_order_value_instead_of_visible_label(self):
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        from src.mac_app import build_app, build_main_window

        app = build_app([])
        window = build_main_window()

        captured = {}

        def fake_run_worker(mode, payload):
            captured["mode"] = mode
            captured["payload"] = payload

        window._run_worker = fake_run_worker
        window.url_input.setText("https://www.youtube.com/watch?v=gtEROmL0NzQ")
        window.api_key_input.setText("AIza-test")
        window.output_dir_input.setText("/tmp")
        window.order_input.setCurrentIndex(0)

        window.export_all()

        self.assertEqual("export_all", captured["mode"])
        self.assertEqual("time", captured["payload"]["order"])
        self.assertNotEqual(window.order_input.currentText(), captured["payload"]["order"])
        window.close()


if __name__ == "__main__":
    unittest.main()
